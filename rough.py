import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import os


"""
  THIS WAS VERSION 1!  

"""
st.title("HRT Calendar Creator")

tab1, tab2 = st.tabs(["By Dose", "By Injection"])
with tab1:
    name = st.text_input("Your Name")
    hrt = st.text_input("Your HRT")
    dose = st.number_input("Your Dose (mg)", min_value=0.0, format="%.2f")
    total_volume = st.number_input("Total Volume of Vial (mL)", min_value=0.0, format="%.2f")
    concentration = st.number_input("Concentration of Vial (mg/mL)", min_value=0.0, format="%.2f")
    interval = st.number_input("Dose Interval (days)", min_value=1, format="%d")
    start_date = st.date_input("Start Date")

    if st.button("Generate"):
        # Workbook creation and setup
        wb = Workbook()
        ws = wb.active
        ws.title = "Calendar"
        ws.sheet_view.showGridLines = False

        # Excel cell styles
        title_fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        font_row = Font(name="Arial", size=12)
        bold_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Worksheet structure and header
        ws['B2'] = f"{name} - {hrt} {dose}mg / {interval} days"
        ws.merge_cells('B2:G2')
        ws['B2'].fill = title_fill
        ws['B2'].font = bold_font
        ws['B2'].alignment = center_align

        ws['A4'] = "Today:"
        ws['B4'] = "=TODAY()"
        ws['B4'].number_format = "DD/MM/YYYY"
        ws['C4'] = total_volume
        ws['D4'] = "mL"
        ws['E4'] = concentration
        ws['F4'] = "mg/mL"

        ws['A5'] = "Dose (mg)"
        ws['B5'] = dose
        ws['C5'] = dose / concentration  # Dose in mL
        ws['D5'] = "mL"

        ws['A6'] = "Interval:"
        ws['B6'] = interval

        ws['A7'] = "Start date"
        ws['B7'] = start_date.strftime("%d/%m/%Y")

        ws['A8'] = "Lasts:"
        ws['B8'] = "=$C$4/($B$5/$E$4)"
        ws['C8'] = "Doses"

        # Table headers
        headers = ["Date", "mL Remaining", "Consecutive days", "Months", "Side", "Nbr. of injections"]
        for col_num, header in enumerate(headers, start=2):
            cell = ws.cell(row=12, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True, size=12)
            cell.alignment = center_align

        # Variable initialization for the data loop
        current_row = 13
        injection_count = 1
        current_side = "Left"
        days_elapsed = 0
        ml_remaining = total_volume

        stream = st.container(height=300)

        # Fill rows with calculated data
        while ml_remaining >= (dose / concentration):
            injection_date = start_date + timedelta(days=days_elapsed)
            dose_ml = dose / concentration

            stream.write(
                f"{injection_count:<12}{injection_date.strftime('%A'):<15}"
                f"{injection_date.strftime('%d/%m/%Y'):<12}{round(ml_remaining, 2):<15}"
                f"{days_elapsed:<15}{round(days_elapsed / 30, 2):<8}{current_side:<10}"
            )

            # Write data to Excel sheet
            ws.cell(row=current_row, column=1, value=injection_date.strftime("%A"))
            ws.cell(row=current_row, column=2, value=injection_date.strftime("%d/%m/%Y"))
            ws.cell(row=current_row, column=3, value=round(ml_remaining, 2))
            ws.cell(row=current_row, column=4, value=days_elapsed)
            ws.cell(row=current_row, column=5, value=round(days_elapsed / 30, 2))
            ws.cell(row=current_row, column=6, value=current_side)
            ws.cell(row=current_row, column=7, value=injection_count)

            # Update variables for the next iteration
            ml_remaining -= dose_ml
            days_elapsed += interval
            current_side = "Right" if current_side == "Left" else "Left"
            injection_count += 1
            current_row += 1

        # Apply thin borders to all cells
        for row in ws.iter_rows(min_row=13, max_row=current_row - 1, min_col=2, max_col=7):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

        # Save the file to Downloads folder
        download_folder = os.path.join(os.path.expanduser("~"), "Downloads")
        output_file = os.path.join(download_folder, f"HRT {name}.xlsx")

        try:
            wb.save(output_file)
            st.success(f"Calendar saved successfully! File saved to: {output_file}")
        except PermissionError:
            st.error("Error: Unable to save the file. Please close any open copies of the file and try again.")
with tab2:
    name2 = st.text_input("Your Name")
    hrt2 = st.text_input("Your HRT")
    injection = st.number_input("Your Injection Amount (ml)", min_value=0.0, format="%.2f")
    total_volume = st.number_input("Total Volume of Vial (mL)", min_value=0.0, format="%.2f")
    interval = st.number_input("Dose Interval (days)", min_value=1, format="%d")
    start_date = st.date_input("Start Date")

    if st.button("Generate"):
        # Workbook creation and setup
        wb = Workbook()
        ws = wb.active
        ws.title = "Calendar"
        ws.sheet_view.showGridLines = False

        # Excel cell styles
        title_fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        font_row = Font(name="Arial", size=12)
        bold_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Worksheet structure and header
        ws['B2'] = f"{name} - {hrt} {dose}mg / {interval} days"
        ws.merge_cells('B2:G2')
        ws['B2'].fill = title_fill
        ws['B2'].font = bold_font
        ws['B2'].alignment = center_align

        ws['A4'] = "Today:"
        ws['B4'] = "=TODAY()"
        ws['B4'].number_format = "DD/MM/YYYY"
        ws['C4'] = total_volume
        ws['D4'] = "mL"
        ws['E4'] = concentration
        ws['F4'] = "mg/mL"

        ws['A5'] = "Dose (mg)"
        ws['B5'] = dose
        ws['C5'] = dose / concentration  # Dose in mL
        ws['D5'] = "mL"

        ws['A6'] = "Interval:"
        ws['B6'] = interval

        ws['A7'] = "Start date"
        ws['B7'] = start_date.strftime("%d/%m/%Y")

        ws['A8'] = "Lasts:"
        ws['B8'] = "=$C$4/($B$5/$E$4)"
        ws['C8'] = "Doses"

        # Table headers
        headers = ["Date", "mL Remaining", "Consecutive days", "Months", "Side", "Nbr. of injections"]
        for col_num, header in enumerate(headers, start=2):
            cell = ws.cell(row=12, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True, size=12)
            cell.alignment = center_align

        # Variable initialization for the data loop
        current_row = 13
        injection_count = 1
        current_side = "Left"
        days_elapsed = 0
        ml_remaining = total_volume

        stream = st.container(height=300)

        # Fill rows with calculated data
        while ml_remaining >= (dose / concentration):
            injection_date = start_date + timedelta(days=days_elapsed)
            dose_ml = dose / concentration

            stream.write(
                f"{injection_count:<12}{injection_date.strftime('%A'):<15}"
                f"{injection_date.strftime('%d/%m/%Y'):<12}{round(ml_remaining, 2):<15}"
                f"{days_elapsed:<15}{round(days_elapsed / 30, 2):<8}{current_side:<10}"
            )

            # Write data to Excel sheet
            ws.cell(row=current_row, column=1, value=injection_date.strftime("%A"))
            ws.cell(row=current_row, column=2, value=injection_date.strftime("%d/%m/%Y"))
            ws.cell(row=current_row, column=3, value=round(ml_remaining, 2))
            ws.cell(row=current_row, column=4, value=days_elapsed)
            ws.cell(row=current_row, column=5, value=round(days_elapsed / 30, 2))
            ws.cell(row=current_row, column=6, value=current_side)
            ws.cell(row=current_row, column=7, value=injection_count)

            # Update variables for the next iteration
            ml_remaining -= dose_ml
            days_elapsed += interval
            current_side = "Right" if current_side == "Left" else "Left"
            injection_count += 1
            current_row += 1

        # Apply thin borders to all cells
        for row in ws.iter_rows(min_row=13, max_row=current_row - 1, min_col=2, max_col=7):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

        # Save the file to Downloads folder
        download_folder = os.path.join(os.path.expanduser("~"), "Downloads")
        output_file = os.path.join(download_folder, f"HRT {name}.xlsx")

        try:
            wb.save(output_file)
            st.success(f"Calendar saved successfully! File saved to: {output_file}")
        except PermissionError:
            st.error("Error: Unable to save the file. Please close any open copies of the file and try again.")

# Run the function to display the Streamlit app
#generate_calendar()