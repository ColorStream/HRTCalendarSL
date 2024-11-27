import PySimpleGUI as sg
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import time


def gera_calendario():

    sg.theme('DarkPurple')

    # Janela principal
    layout = [  
        [sg.Text('HRT Calendar creator')],
        [sg.Text('Your name: '), sg.InputText(key='nome')],
        [sg.Text('Your ester (EV, EEn etc...): '), sg.InputText(key='ester')],
        [sg.Text('Your dose (mg): '), sg.InputText(key='dose')],
        [sg.Text('Your vial total volume (mL): '), sg.InputText(key='vol_total')],
        [sg.Text('Concentration of your vial (mg/mL): '), sg.InputText(key='concentracao')],
        [sg.Text('Dose interval (days): '), sg.InputText(key='intervalo')],
        [sg.Text('Start date (dd/mm/yyyy): '), sg.InputText(key='inicio')],
        [sg.Button('Generate'), sg.Button('Clean'), sg.Button('Cancel')],
        [sg.Output(s=(75, 10))]
    ]

    window = sg.Window('HRT Calendar generator', layout, font=("Bookman Old Style", 16))
         
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == 'Cancel':
            break
        if event == 'Clean':
            for key in values:
                window[key]('')

        if event == 'Generate':
            try:
                # Entrada do usuário
                nome_usr = values['nome']
                ester = values['ester']
                dose = float(values['dose'])
                ml_total = float(values['vol_total'])
                concentracao = float(values['concentracao'])
                intervalo = int(values['intervalo'])
                data_inicio = datetime.strptime(values['inicio'], "%d/%m/%Y")
            except ValueError:
                sg.popup("Error: verify the values. Make sure to follow the DD/MM/YYY format.")
                continue

            # Criar planilha
            wb = Workbook()
            ws = wb.active
            ws.title = "Calendar"
            ws.sheet_view.showGridLines = False

            # Estilos para o Excel
            titulo_fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            font_row = Font(name="Bookman Old Style", size=12)
            bold_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            # Estrutura da planilha
            ws['B2'] = f"{nome_usr} - {ester} {dose}mg / {intervalo} days"
            ws.merge_cells('B2:G2')
            ws['B2'].fill = titulo_fill
            ws['B2'].font = bold_font
            ws['B2'].alignment = center_align

            # Primeiras linhas
            ws['A4'] = "Today:"
            ws['B4'] = "=TODAY()"
            ws['B4'].number_format = "DD/MM/YYYY"
            ws['C4'] = ml_total
            ws['D4'] = "mL"
            ws['E4'] = concentracao
            ws['F4'] = "mg/mL"

            ws['A5'] = "Dose (mg)"
            ws['B5'] = dose
            ws['C5'] = dose / concentracao  # Cálculo da dose em mL
            ws['D5'] = "mL"

            ws['A6'] = "Interval:"
            ws['B6'] = intervalo

            ws['A7'] = "Start date"
            ws['B7'] = data_inicio.strftime("%d/%m/%Y")

            ws['A8'] = "Lasts:"
            ws['B8'] = "=$C$4/($B$5/$E$4)"
            ws['C8'] = "Doses"

            # Cabeçalhos
            headers = ["Date", "mL Remaining", "Consecutive days", "Months", "Side", "Nbr. of injections"]
            for col_num, header in enumerate(headers, start=2):
                cell = ws.cell(row=12, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True, size=12)
                cell.alignment = center_align

            # Variáveis de controle
            linha_atual = 13
            aplicacao = 1
            lado_atual = "Left"
            dias_corridos = 0
            ml_restante = ml_total

            # Loop para preencher as linhas
            while ml_restante >= (dose / concentracao):
                data_aplicacao = data_inicio + timedelta(days=dias_corridos)
                ml_dose = dose / concentracao

                print(
                f"{aplicacao:<12}{data_aplicacao.strftime('%A'):<15}"
                f"{data_aplicacao.strftime('%d/%m/%Y'):<12}{round(ml_restante, 2):<15}"
                f"{dias_corridos:<15}{round(dias_corridos / 30, 2):<8}{lado_atual:<10}"
                )

                # Preencher os dados na planilha
                ws.cell(row=linha_atual, column=1, value=data_aplicacao.strftime("%A"))  # Dia
                ws.cell(row=linha_atual, column=2, value=data_aplicacao.strftime("%d/%m/%Y"))  # Data
                ws.cell(row=linha_atual, column=3, value=round(ml_restante, 2))  # mL restante
                ws.cell(row=linha_atual, column=4, value=dias_corridos)  # Dias corridos
                ws.cell(row=linha_atual, column=5, value=round(dias_corridos / 30, 2))  # Meses
                ws.cell(row=linha_atual, column=6, value=lado_atual)  # Lado
                ws.cell(row=linha_atual, column=7, value=aplicacao)  # Aplicações

                # Atualizar variáveis
                ml_restante -= ml_dose
                dias_corridos += intervalo
                lado_atual = "Right" if lado_atual == "Left" else "Right"
                aplicacao += 1
                linha_atual += 1

            # Definir borda fina para todas as células
                thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
                )

            # Definir borda inferior para as células
            thin_border_bottom = Border(
                bottom=Side(style="thin")
            )

            # Aplicar bordas apenas às células de dados
            for row in ws.iter_rows(min_row=13, max_row=linha_atual - 1, min_col=2, max_col=7):
                for cell in row:
                    cell.border = thin_border_bottom
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # ALinhamento central

            for row in ws.iter_rows(min_row=1, max_row=linha_atual - 1, min_col=1, max_col=7):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajustar larguras das colunas
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max_length + 2

            # Salvar o arquivo
            pasta_downloads = os.path.join(os.path.expanduser("~"), "Downloads")
            arquivo_saida = os.path.join(pasta_downloads, f"HRT {nome_usr}.xlsx")

            try:
                wb.save(arquivo_saida)
                sg.popup(f"Calendar saved successfuly!\nFile saved to: {arquivo_saida}")
            except PermissionError:
                sg.popup("ERROR. I'm too lazy to translate this right now.")
    
gera_calendario()
