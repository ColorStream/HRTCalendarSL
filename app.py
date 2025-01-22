# INITIAL IMPORTS
import streamlit as st
from datetime import timedelta, date
import pandas as pd

class Form:
    """ A class for generating an form object in the homepage.
    """
    def __init__(self, name=None, ester=None, interval=None, start_date=None, dose=None, dosetype=None, total_volume=None, concentration=None):
        """The Form object's attributes for generating a calendar.

        Args:
            name (str, optional): The user's name. Defaults to None.
            ester (str, optional): The user's ester type. Defaults to None.
            interval (int): The interval in which shots need to be delivered. Defaults to None.
            start_date (str): The start date in which to generate from. Defaults to None.
            dose (int): The user's dosage amount. Defaults to None.
            dosetype (str): The dose type for generating the form (mg/mL). Defaults to None.
            total_volume (int): The total volume of a user's hormone vial. Defaults to None.
            concentration (int): The user's hormone vial concentration (mg/mL). Defaults to None.
        """
        self.name = name
        self.ester = ester
        self.interval = interval
        self.start_date = start_date
        self.dose = dose
        self.dosetype = dosetype
        self.total_volume = total_volume
        self.concentration = concentration

    def create_inputs(self, date_format):
        """Creates the inputs for setting the form object's values.

        Args:
            date_format (str): The date format for generating the input form. 
        """
        self.name = st.text_input("Your Name", value=self.name)
        self.ester = st.text_input("Your [Ester](https://en.wikipedia.org/wiki/Steroid_ester)", value=self.ester)
        self.interval = st.number_input("Dose Interval (days)", min_value=1, format="%d", value=self.interval)
        self.start_date = st.date_input("Start Date", value=self.start_date, format=date_format)

        # dosage inputs
        self.dose = st.number_input(f"Your Dose ({self.dosetype})", min_value=0.0, format="%.2f", value=self.dose)
        self.total_volume = st.number_input("Total Volume of Vial (mL)", min_value=0.0, format="%.2f", value=self.total_volume)
        self.concentration = st.number_input("Concentration of Vial (mg/mL)", min_value=0.0, format="%.2f", value=self.concentration)

    def generate_metadata(self, date_format, injections_sum):
        """Generates a metadata dataframe for the full calendar generation method.

        Args:
            date_format (str): The user's preferred date format.
            injections_sum (int): The amount of doses that a vial can contain.

        Returns:
            metadata_df (dataframe): A pandas dataframe containing the calendar's contextual information.
        """
        if self.dosetype == "mg":
            dosemg = self.dose
            doseml = self.dose / self.concentration
        else:
            dosemg = self.dose * self.concentration
            doseml = self.dose
        metadata = {
            "Name": self.name,
            "Ester": self.ester,
            "Generation Date": f"{date.today().strftime(date_format)}",
            "Dose (mg)": f"{dosemg} mg",
            "Dose (mL)": f"{doseml} mL",
            "Interval": self.interval,
            "Start Date": self.start_date.strftime(date_format),
            "Lasts": f"{injections_sum} doses",
            "Total mL": f"{self.total_volume} mL",
            "Concentration (mg/mL)": f"{self.concentration} mg/mL"
        }
        metadata_df = pd.DataFrame([metadata])
        return metadata_df
    
    def validate_fields(self):
        """Validates whether the required information is there.

        Returns:
            bool: True if required rields are filled out, False if they aren't.
        """
        required_fields = [self.interval, self.start_date, self.dose, self.total_volume, self.concentration]
        return all(field is not None and field != '' for field in required_fields)

    def generate_calendar(self, date_format, starting_side):
        """Generates two dataframes with general information and an injection schedule.

        Args:
            date_format (str): The date format used for the schedule.
            starting_side (str): The starting side input by the user.

        Returns:
            metadata, df (tuple): The metadata and calendar dataframes in a tuple.
        """
        # dictionary, since you can't append to a dataframe
        data_dict = {
            "Injection Count": [],
            "Day": [],
            "Date": [],
            "Remaining mL": [],
            "Days": [],
            "Months": [],
            "Side": []
        }

        # variable initialization
        injection_count = 1
        current_side = starting_side
        days_elapsed = 0
        ml_remaining = self.total_volume
        if self.dosetype == "mg":
            dose_ml = self.dose / self.concentration
        else:
            dose_ml = self.dose

        # loop to generate the calendar data and store it in the dictionary
        while ml_remaining >= dose_ml:
            injection_date = self.start_date + timedelta(days=days_elapsed)

            # append data to the dictionary
            data_dict["Injection Count"].append(injection_count)
            data_dict["Day"].append(injection_date.strftime("%A"))
            data_dict["Date"].append(injection_date.strftime(date_format))
            data_dict["Remaining mL"].append(round(ml_remaining, 2))
            data_dict["Days"].append(days_elapsed)
            data_dict["Months"].append(round(days_elapsed / 30, 2))
            data_dict["Side"].append(current_side)

            # update variables for the next iteration
            ml_remaining -= dose_ml
            days_elapsed += self.interval
            current_side = "Right" if current_side == "Left" else "Left"
            injection_count += 1

        #convert the dictionary to a DataFrame and display
        metadata = self.generate_metadata(date_format, injection_count)
        df = pd.DataFrame(data_dict)
        return metadata, df


# OPENPYXL IMPORTS
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io

def export_to_excel(metadata_df, calendar_df):
    """Writes the metadata and calendar dataframe to an excel sheet and returns it.

    Args:
        metadata_df (dataframe): A dataframe containing the user's name and full dosage information.
        calendar_df (dataframe): A dataframe containing the injection schedule based on the metadata.

    Returns:
        output (BytesIO object): The .xlsx file using an in-memory bytes buffer.
    """
    if metadata_df is None or calendar_df is None:
        st.error("Metadata or Calendar is not initialized.")
        return

    wb = Workbook()
    ws = wb.active

    try:
        output = io.BytesIO()
        # write metadata
        for r_idx, row in enumerate(dataframe_to_rows(metadata_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        md_fill = PatternFill(start_color='cce6ff', end_color='cce6ff', fill_type='solid') # blue :)
        for cell in ws[1]: #color metadata header
            cell.fill = md_fill

        # write calendar data
        calendar_start_row = len(metadata_df) + 3  # leave 2 rows gap between metadata and calendar
        for r_idx, row in enumerate(dataframe_to_rows(calendar_df, index=False, header=True), calendar_start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        cal_fill = PatternFill(start_color='ffccff', end_color='ffccff', fill_type='solid') # pink :)
        for cell in ws[calendar_start_row]: #color calendar header
            cell.fill = cal_fill
        
        # this method is taken from here: https://github.com/Loke-60000/Archaeological-Sites-Sorter/blob/master/app.py
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # get the column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2  # padding
            ws.column_dimensions[column].width = adjusted_width

        # save to downloads
        wb.save(output)
        output.seek(0)  # Move the cursor to the start of the stream
        
        return output
    
    except Exception as e:
        return f"An unexpected error occurred: {e}"

# ICALENDAR EXPORTS
from icalendar import Calendar, Event
from datetime import datetime

def export_to_ical(calendar_df, warn_secondtolast=False):
    """Exports the generated injection calendar to a .ics file using the icalendar module. 

    Args:
        calendar_df (dataframe): The pandas dataframe generated from the form.generate_calendar() method.
        warn_secondtolast (bool, optional): Whether the user wants a warning on the second to last dose. Defaults to False.

    Returns:
        output (IOBase object): The .ical file using an in-memory bytes buffer.
    """
    if calendar_df is None:
        st.error("Calendar data is not provided.")
        return

    try:
        output = io.BytesIO() 
        cal = Calendar()
        cal.add('prodid', '-//HRT Calendar//mxm.dk//')
        cal.add('version', '2.0')

        # get the index of the second-to-last dose if the warning is enabled
        second_to_last_index = len(calendar_df) - 2 if warn_secondtolast and len(calendar_df) > 1 else None

        for index, row in calendar_df.iterrows():
            event = Event()

            # default event name
            event.add('summary', f"Take dose: {row['Side']} side")

            # if this is the second-to-last dose, add a special message
            if warn_secondtolast and index == second_to_last_index:
                event['summary'] = event['summary'] + " (SECOND TO LAST DOSE)"
                event.add('description', "Please order a refill if you haven't already!\n")
            else:
                event.add('description', "")

            description = (f"Injection Count: {row['Injection Count']}\n"
                           f"Remaining mL: {row['Remaining mL']}\n"
                           f"Days: {row['Days']}\n"
                           f"Months: {row['Months']}\n")
            event['description'] = event['description'] + description

            # set the event date
            event_date = datetime.strptime(row['Date'], "%Y-%m-%d")  
            event.add('dtstart', event_date.date())
            event.add('dtend', event_date.date())  # all-day event, so start and end are the same

            cal.add_component(event)

        # save to downloads folder
        output = io.BytesIO()
        output.write(cal.to_ical())
        output.seek(0)  # move the cursor to the start of the stream

        return output

    except Exception as e:
        return f"An unexpected error occurred: {e}"

def get_date_format(date_format):
    """Given a selected date format, returns a proper date format for automatically formatting dates in the calendar.

    Args:
        date_format (str): Conventional date formats input by the user.

    Returns:
        format (str): Returns the proper string for formatting dates.
    """
    if date_format == "DD/MM/YYYY":
        format = "%d/%m/%Y"
    elif date_format == "MM/DD/YYYY":
        format = "%m/%d/%Y"
    else:
        format = "%Y/%m/%d"
    return format

def generate_form(dosetype, date_format, starting_side, warn_secondtolast):
    """Generate the form object on the homepage according to dosetype selected in the sidebar. 

    Args:
        dosetype (str): The dosetype ("mg", "mL") selected by the user.
        date_format (str): The date format that the dataframe/excel sheet will export.
        starting_side (str): The starting side ("Left", "Right") selected by the user.
        warn_secondtolast (str): Whether the user wants a warning on the second to last dose in the iCal file.
    """
    st.title(f"Calendar By Dosage in {dosetype}")
    form = Form(dosetype=dosetype)
    form.create_inputs(date_format=date_format)

    # All the stuff I have to initialize....... I feel like this is bad practice somehow...
    invalidForm = False
    dataframe_display = ""
    exceloutput = ""
    icaloutput = ""

    col1_, col2_, col3_ = st.columns([2,1,2])
    with col1_:
        if st.button(f"Generate Dosage ({dosetype}) Calendar"):
            if not form.validate_fields():
                invalidForm = True
            else:
                date_format = get_date_format(date_format)
                metadata, calendar = form.generate_calendar(date_format=date_format, starting_side=starting_side)
                dataframe_display = [metadata, calendar]
    with col2_:
        if st.button("Export to Excel"):
            if not form.validate_fields():
                invalidForm = True
            else:
                date_format = get_date_format(date_format)
                metadata, calendar = form.generate_calendar(date_format=date_format, starting_side=starting_side)
                exceloutput = export_to_excel(metadata_df=metadata, calendar_df=calendar)
                
    with col3_:
        if st.button("Export to iCal"):
            if not form.validate_fields():
                invalidForm = True
            else:
                warn_secondtolast = False if warn_secondtolast == "No" else True
                metadata, calendar = form.generate_calendar(date_format="%Y-%m-%d", starting_side=starting_side)
                icaloutput = export_to_ical(calendar_df=calendar, warn_secondtolast=warn_secondtolast)

    # moving things below the columns because it's ugly within them
    if invalidForm:
        st.error("Please fill out all fields properly before exporting the calendar.")
    
    if dataframe_display:
        st.dataframe(dataframe_display[0])
        st.dataframe(dataframe_display[1])
    
    if exceloutput:
        if "Error" in exceloutput or "unexpected" in exceloutput:
            st.error(exceloutput)  
        else:
            st.success("File generated.") 
            st.download_button(
            label="Download Excel",
            data=exceloutput,
            file_name=f"{form.name}_hrt_calendar.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    if icaloutput:
        if "Error" in icaloutput or "unexpected" in icaloutput:
            st.error(icaloutput)
        else:
            st.success("File generated.") 
            st.download_button(
            label="Download iCal",
            data=icaloutput,
            file_name=f"{form.name}_hrt_calendar.ics",
            mime="text/calendar"
            )

def main():
    """
    Runs the main page for users to see.
    """
    st.set_page_config(page_title="HRT Calendar Generator", page_icon=None, layout="centered", initial_sidebar_state="auto", menu_items=None)

    # markdown for the button columns
    st.markdown("""
            <style>
                div[data-testid="column"] {
                    width: fit-content !important;
                    flex: unset;
                }
                div[data-testid="column"] * {
                    width: fit-content !important;
                }
            </style>
            """, unsafe_allow_html=True)

    st.sidebar.title("HRT Calendar Generator")
    tab = st.sidebar.radio("Choose Calculation Method", ["By dosage in mg", "By dosage in mL"])
    date_format = st.sidebar.radio("Choose Date Format", ["MM/DD/YYYY", "DD/MM/YYYY", "YYYY/MM/DD"]) #apologies for americanizing this...
    starting_side = st.sidebar.radio("Choose Starting Side", ["Left", "Right"])
    warn_secondtolast = st.sidebar.radio("Warning at Second to Last Dose? (For iCal)", ["No", "Yes"])
    st.sidebar.write("Newbie to injectable hormones? [Check out this resource.](https://www.folxhealth.com/library/hormone-self-injection-guide)") 
    st.sidebar.divider()
    st.sidebar.caption("[Github Repo](https://github.com/ColorStream/HRTCalendarSL)")

    if tab == "By dosage in mg":
        generate_form(dosetype="mg", date_format=date_format, starting_side=starting_side, warn_secondtolast=warn_secondtolast)

    elif tab == "By dosage in mL":
        generate_form(dosetype="mL", date_format=date_format, starting_side=starting_side, warn_secondtolast=warn_secondtolast)

main()
