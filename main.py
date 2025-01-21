import pandas as pd
import streamlit as st
from datetime import timedelta, date
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pandas as pd
import os

class Form:
    def __init__(self, name=None, hrt=None, interval=None, start_date=None, dose=None, dosetype=["mg", "mL"], total_volume=None, concentration=None):
        self.name = name
        self.hrt = hrt
        self.interval = interval
        self.start_date = start_date
        self.dose = dose
        self.dosetype = dosetype
        self.total_volume = total_volume
        self.concentration = concentration

    def create_inputs(self, date_format):
        self.name = st.text_input("Your Name", value=self.name)
        self.hrt = st.text_input("Your HRT", value=self.hrt)
        self.interval = st.number_input("Dose Interval (days)", min_value=1, format="%d", value=self.interval)
        self.start_date = st.date_input("Start Date", value=self.start_date, format=date_format)

        # dosage inputs
        self.dose = st.number_input(f"Your Dose ({self.dosetype})", min_value=0.0, format="%.2f", value=self.dose)
        self.total_volume = st.number_input("Total Volume of Vial (mL)", min_value=0.0, format="%.2f", value=self.total_volume)
        self.concentration = st.number_input("Concentration of Vial (mg/mL)", min_value=0.0, format="%.2f", value=self.concentration)

    def generate_metadata(self, date_format, injections_sum):
        if self.dosetype == "mg":
            dosemg = self.dose
            doseml = self.dose / self.concentration
        else:
            dosemg = self.dose * self.concentration
            doseml = self.dose
        metadata = {
            "Name": self.name,
            "HRT Type": self.hrt,
            "Generation Date": f"{date.today().strftime(date_format)}",
            "Dose (mg)": f"{dosemg} mg",
            "Dose (mL)": f"{doseml} mL",
            "Interval": self.interval,
            "Start Date": self.start_date.strftime(date_format),
            "Lasts": f"{injections_sum} doses",
            "Total mL": f"{self.total_volume} mL",
            "Concentration (mg/mL)": f"{self.concentration} mg/mL"
        }
        metadata_df = pd.DataFrame([metadata])
        return metadata_df
    
    def validate_fields(self):
        required_fields = [self.name, self.hrt, self.interval, self.start_date, self.dose, self.total_volume, self.concentration]
        return all(field is not None and field != '' for field in required_fields)

    def generate_calendar(self, date_format, starting_side):
        # dictionary to store data before displaying as dataframe
        data_dict = {
            "Injection Count": [],
            "Day": [],
            "Date": [],
            "Remaining mL": [],
            "Days": [],
            "Months": [],
            "Side": []
        }

        # variable initialization
        injection_count = 1
        current_side = starting_side
        days_elapsed = 0
        ml_remaining = self.total_volume
        if self.dosetype == "mg":
            dose_ml = self.dose / self.concentration
        else:
            dose_ml = self.dose

        # loop to generate the calendar data and store it in the dictionary
        while ml_remaining >= dose_ml:
            injection_date = self.start_date + timedelta(days=days_elapsed)

            # append data to the dictionary
            data_dict["Injection Count"].append(injection_count)
            data_dict["Day"].append(injection_date.strftime("%A"))
            data_dict["Date"].append(injection_date.strftime(date_format))
            data_dict["Remaining mL"].append(round(ml_remaining, 2))
            data_dict["Days"].append(days_elapsed)
            data_dict["Months"].append(round(days_elapsed / 30, 2))
            data_dict["Side"].append(current_side)

            # update variables for the next iteration
            ml_remaining -= dose_ml
            days_elapsed += self.interval
            current_side = "Right" if current_side == "Left" else "Left"
            injection_count += 1

        #convert the dictionary to a DataFrame and display
        metadata = self.generate_metadata(date_format, injection_count)
        df = pd.DataFrame(data_dict)
        return metadata, df

def export_to_excel(metadata_df, calendar_df, filename="hrt_calendar.xlsx"):
        if metadata_df is None or calendar_df is None:
            st.error("Metadata or Calendar is not initialized.")
            return

        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active

        try:
            # Write metadata to Excel
            for r_idx, row in enumerate(dataframe_to_rows(metadata_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for cell in ws[1]: #color metadata header
                cell.fill = header_fill

            # Write calendar data to Excel
            calendar_start_row = len(metadata_df) + 3  # Leave 2 rows gap between metadata and calendar
            for r_idx, row in enumerate(dataframe_to_rows(calendar_df, index=False, header=True), calendar_start_row):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            for cell in ws[calendar_start_row]: #color calendar header
                cell.fill = header_fill
            
            # This method is taken from here: https://github.com/Loke-60000/Archaeological-Sites-Sorter/blob/master/app.py
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2  # Add some padding
                ws.column_dimensions[column].width = adjusted_width

            # Save the file to Downloads folder
            download_folder = os.path.join(os.path.expanduser("~"), "Downloads")
            output_file = os.path.join(download_folder, filename)

            wb.save(output_file)
            st.success(f"Calendar saved successfully! File saved to: {output_file}")

        except PermissionError:
            st.error("Error: Unable to save the file. Please close any open copies of the file and try again.")
        except Exception as e:
            st.error(f"An unexpected error occurred: {e}")

def get_date_format(date_format):
    if date_format == "DD/MM/YYYY":
        format = "%d/%m/%Y"
    elif date_format == "MM/DD/YYYY":
        format = "%m/%d/%Y"
    else:
        format = "%Y/%m/%d"
    return format

def main():
    st.sidebar.title("HRT Calendar Generator")
    tab = st.sidebar.radio("Choose Calculation Method", ["By dosage in mg", "By dosage in mL"])
    date_format = st.sidebar.radio("Choose Date Format", ["MM/DD/YYYY", "DD/MM/YYYY", "YYYY/MM/DD"]) #apologies for americanizing this. 
    starting_side = st.sidebar.radio("Choose Starting Side", ["Left", "Right"])

    if tab == "By dosage in mg":
        st.title("Calendar By Dosage in mg")
        form = Form(dosetype="mg")
        form.create_inputs(date_format=date_format)
        if st.button("Generate Dosage (mg) Calendar"):
            if not form.validate_fields():
                st.warning("Please fill out all fields properly before generating the calendar.")
            else:
                date_format = get_date_format(date_format)
                metadata, calendar = form.generate_calendar(date_format=date_format, starting_side=starting_side)
                st.dataframe(metadata)
                st.dataframe(calendar)
                
        if st.button("Export to Excel"):
            if not form.validate_fields():
                st.warning("Please fill out all fields properly before generating the calendar.")
            else:
                date_format = get_date_format(date_format)
                metadata, calendar = form.generate_calendar(date_format=date_format, starting_side=starting_side)
                export_to_excel(metadata_df=metadata, calendar_df=calendar)
        

    elif tab == "By dosage in mL":
        st.title("Calendar By Dosage in mL")
        form = Form(dosetype="mL")
        form.create_inputs(date_format=date_format)
        if st.button("Generate Dosage (mL) Calendar"):
            if not form.validate_fields():
                st.warning("Please fill out all fields properly before trying to generate the calendar.")
            else:
                date_format = get_date_format(date_format)
                metadata, calendar = form.generate_calendar(date_format=date_format, starting_side=starting_side)
                st.dataframe(metadata)
                st.dataframe(calendar)

        if st.button("Export to Excel"):
            if not form.validate_fields():
                st.warning("Please fill out all fields properly before generating the calendar.")
            else:
                date_format = get_date_format(date_format)
                metadata, calendar = form.generate_calendar(date_format=date_format, starting_side=starting_side)
                export_to_excel(metadata_df=metadata, calendar_df=calendar)

# Run the app
main()
